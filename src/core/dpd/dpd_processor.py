import logging
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

class DPDProcessor:
  def __init__(self):
    self.logger = logging.getLogger(__name__)

    # DPD模板工作表映射关系（数据字段 -> 模板字段）
    self.sheet_mappings = {
      "List （运单清单）": {
        "客户单号": "Remark\n（箱唛 or  FBA ）",
        "转单号": "Tracking No",
        "国家二字码": "County",
        "件数": "PCS",
        "收货实重": "GW (kg)",
        "收货材积重": "VW（kg)",
        "方数": "Cubic Number(CBM)",
        "收件人邮编": "post Code"
      },
      "总结单": {
        "de_postcodes": ["4347", "6126", "14656", "21423", "36251", "39171", "44145", "47495", "56068", "59368", "67227", "75177", "90451"],
        "supported_countries": ["DE", "FR", "IT", "ES", "NL", "PL", "CZ", "BE"],
        "de_start_col": 6,  # F列
        "other_countries_start_col": 19,  # S列
        "other_col": 26,  # Z列
        "total_col": 27,  # AA列
        "data_row": 4  # 数据填充行
      },
      "子单号": {
        "detail": {
          "客户单号": "参考号 （必填）",
          "子转单号": "子单号（必填）",
        },
        "list": {
          "转单号": "主单号（必填）",
          "收件人公司": "公司",
          "收件人姓名": "收件人",
          "方数": "方数"
        }
      }
    }

  def get_template_workbook(self, template_path: str):
    """
    获取模板工作簿
    """
    try:
        if not Path(template_path).exists():
            self.logger.error(f"模板文件不存在: {template_path}")
            return None
        return load_workbook(template_path)
    except FileNotFoundError:
        self.logger.error(f"模板文件未找到: {template_path}")
        return None
    except PermissionError:
        self.logger.error(f"模板文件访问权限不足: {template_path}")
        return None
    except Exception as e:
        self.logger.error(f"获取模板工作簿时出错: {str(e)}")
        return None

  def process_dpd_data(self, original_file_data: pd.DataFrame, original_detail_file_data: pd.DataFrame, template_path: str, output_path: str):
    """
    处理DPD数据并填充到模板中

    Args:
        original_file_data (pd.DataFrame): 原始文件数据
        original_detail_file_data (pd.DataFrame): 原始明细表数据
        template_path (str): DPD模板路径
        output_path (str): 输出文件路径

    Returns:
        bool: 处理结果
    """
    try:
        self.logger.info("开始处理DPD数据")

        template_workbook = self.get_template_workbook(template_path)
        if template_workbook is None:
            return False

        original_file_data_count = original_file_data.shape[0]

        # 处理运单清单工作表
        list_sheet, first_empty_row, collection_total_row = self.get_template_list_sheet(template_workbook)
        if list_sheet is not None:
            self.process_list_sheet(template_workbook, list_sheet, original_file_data, first_empty_row, collection_total_row, original_file_data_count)

        # 处理子单号工作表
        sub_order_sheet, first_empty_row, collection_total_row = self.get_template_sub_order_sheet(template_workbook)
        if sub_order_sheet is not None:
            self.process_sub_order_sheet(template_workbook, sub_order_sheet, original_detail_file_data, original_file_data, first_empty_row, collection_total_row, original_file_data_count)

        # 处理总结单工作表
        summary_sheet = self.get_template_summary_sheet(template_workbook)
        if summary_sheet is not None:
            self.process_summary_sheet(template_workbook, summary_sheet, original_file_data)

        # 保存文件
        template_workbook.save(output_path)
        self.logger.info(f"DPD数据处理完成，输出文件: {output_path}")
        return True
    except Exception as e:
        self.logger.error(f"处理DPD数据时出错: {str(e)}")
        return False

  def get_header_column_mapping(self, worksheet, header_row=1):
    """
    获取模板表头到列号的映射关系

    Args:
        worksheet: openpyxl工作表对象
        header_row: 表头所在行号（默认第1行）

    Returns:
        dict: 表头名称到列号的映射字典
    """
    header_mapping = {}
    try:
        max_col = worksheet.max_column
        self.logger.info(f"开始分析表头，共{max_col}列")

        for col in range(1, max_col + 1):
            header_value = worksheet.cell(row=header_row, column=col).value
            if header_value is not None:
                header_name = str(header_value).strip()
                if header_name:  # 非空表头
                    header_mapping[header_name] = col
                    self.logger.debug(f"表头映射: '{header_name}' -> 列{col}")

        self.logger.info(f"表头映射完成，共找到{len(header_mapping)}个有效表头")
        self.logger.debug(f"表头映射详情: {header_mapping}")

    except Exception as e:
        self.logger.error(f"分析表头时出错: {str(e)}")

    return header_mapping

  def get_template_list_sheet(self, template_workbook: Workbook):
    """
    获取模板中的运单清单工作表

    Args:
        template_workbook: 模板工作簿

    Returns:
        tuple: (worksheet, first_empty_row, collection_total_row)
    """
    try:
        self.logger.info("开始获取模板运单清单工作表")
        sheet_names = template_workbook.sheetnames
        self.logger.info(f"模板包含的工作表: {sheet_names}")

        # 尝试多种可能的运单清单工作表名称
        possible_names = ["List （运单清单）", "List", "运单清单", "List（运单清单）", "运单清单 （List）"]
        list_sheet = None
        for name in possible_names:
            if name in sheet_names:
                list_sheet = template_workbook[name]
                self.logger.info(f"找到运单清单工作表: '{name}'")
                break

        if list_sheet is None:
            self.logger.error("未找到运单清单工作表")
            return None, None, None

        max_row = list_sheet.max_row
        max_col = list_sheet.max_column
        self.logger.info(f"运单清单工作表范围: {max_row}行 x {max_col}列")

        first_empty_row = self._find_first_empty_row(list_sheet, max_row, max_col)
        collection_total_row = self._find_collection_total_row(list_sheet, max_row, max_col, 3)

        self.logger.info(f"运单清单工作表分析完成:")
        self.logger.info(f"  - 工作表名称: '{list_sheet.title}'")
        self.logger.info(f"  - 第一个空行: 第{first_empty_row}行")
        self.logger.info(f"  - Collection Total行: 第{collection_total_row}行" if collection_total_row else "  - Collection Total行: 未找到")

        return list_sheet, first_empty_row, collection_total_row

    except Exception as e:
        self.logger.error(f"获取运单清单工作表时出错: {str(e)}")
        return None, None, None

  def process_list_sheet(self, template_workbook: Workbook, list_sheet: Worksheet, original_file_data: pd.DataFrame, first_empty_row: int, collection_total_row: int, original_file_data_count: int):
    """
    处理运单清单工作表
    """
    if template_workbook is None or list_sheet is None:
        self.logger.error("获取模板工作表失败")
        return False

    try:
        self.fill_list_sheet(list_sheet, original_file_data, first_empty_row)
        self.logger.info("运单清单工作表处理完成")
    except Exception as e:
        self.logger.error(f"处理运单清单工作表时出错: {str(e)}")
        raise

  def fill_list_sheet(self, list_sheet: Worksheet, original_file_data: pd.DataFrame, first_empty_row: int):
    """
    填充运单清单工作表
    """
    try:
        # 获取数据字段到模板字段的映射关系
        field_mappings = self.sheet_mappings["List （运单清单）"]

        # 获取模板表头到列号的映射
        header_column_mapping = self.get_header_column_mapping(list_sheet)

        # 检查原始数据的列名和可用数据
        available_columns = original_file_data.columns.tolist()
        self.logger.info(f"原始数据可用列: {available_columns}")
        self.logger.info(f"模板表头到列号映射: {header_column_mapping}")

        # 数据行数
        data_row_count = len(original_file_data)
        self.logger.info(f"需要填充 {data_row_count} 行数据")

        # 逐行填充数据
        for data_row_idx in range(data_row_count):
            target_row = first_empty_row + data_row_idx
            row_data = original_file_data.iloc[data_row_idx]

            self.logger.debug(f"开始填充第 {data_row_idx + 1} 行数据到模板第 {target_row} 行")

            # 遍历字段映射关系
            for data_field, template_field in field_mappings.items():
                try:
                    # 检查数据中是否有该字段
                    if data_field in available_columns:
                        # 检查模板中是否有对应的表头列
                        if template_field in header_column_mapping:
                            col_num = header_column_mapping[template_field]
                            field_value = row_data[data_field]

                            # 填充数据到指定位置
                            list_sheet.cell(row=target_row, column=col_num).value = field_value
                            self.logger.debug(f"  填充: {data_field}({field_value}) -> {template_field}(列{col_num})")
                        else:
                            self.logger.warning(f"  模板中未找到表头 '{template_field}'，跳过字段 '{data_field}'")
                    else:
                        self.logger.warning(f"  原始数据中未找到字段 '{data_field}'，跳过")

                except Exception as e:
                    self.logger.error(f"  填充字段 '{data_field}' 时出错: {str(e)}")
                    continue

        self.logger.info(f"运单清单数据填充完成，共填充 {data_row_count} 行")

    except Exception as e:
        self.logger.error(f"填充运单清单工作表时出错: {str(e)}")
        raise

  def get_template_sub_order_sheet(self, template_workbook: Workbook):
    """
    获取模板中的子单号工作表

    Args:
        template_workbook: 模板工作簿

    Returns:
        tuple: (worksheet, first_empty_row, collection_total_row)
    """
    try:
        self.logger.info("开始获取模板子单号工作表")
        sheet_names = template_workbook.sheetnames
        self.logger.info(f"模板包含的工作表: {sheet_names}")

        # 尝试多种可能的子单号工作表名称
        possible_names = ["子单号", "Sub Order Number", "Sub Order", "子单号信息"]
        sub_order_sheet = None
        for name in possible_names:
            if name in sheet_names:
                sub_order_sheet = template_workbook[name]
                self.logger.info(f"找到子单号工作表: '{name}'")
                break

        if sub_order_sheet is None:
            self.logger.error("未找到子单号工作表")
            return None, None, None

        max_row = sub_order_sheet.max_row
        max_col = sub_order_sheet.max_column
        self.logger.info(f"子单号工作表范围: {max_row}行 x {max_col}列")

        first_empty_row = self._find_first_empty_row(sub_order_sheet, max_row, max_col)
        collection_total_row = self._find_collection_total_row(sub_order_sheet, max_row, max_col)

        self.logger.info(f"子单号工作表分析完成:")
        self.logger.info(f"  - 工作表名称: '{sub_order_sheet.title}'")
        self.logger.info(f"  - 第一个空行: 第{first_empty_row}行")
        self.logger.info(f"  - Collection Total行: 第{collection_total_row}行" if collection_total_row else "  - Collection Total行: 未找到")

        return sub_order_sheet, first_empty_row, collection_total_row

    except Exception as e:
        self.logger.error(f"获取子单号工作表时出错: {str(e)}")
        return None, None, None

  def process_sub_order_sheet(self, template_workbook: Workbook, sub_order_sheet: Worksheet, original_detail_file_data: pd.DataFrame, original_file_data: pd.DataFrame, first_empty_row: int, collection_total_row: int, original_file_data_count: int):
    """
    处理子单号工作表
    """
    if template_workbook is None or sub_order_sheet is None:
        self.logger.error("获取模板工作表失败")
        return False

    try:
        self.fill_sub_order_sheet(sub_order_sheet, original_detail_file_data, original_file_data, first_empty_row)
        self.logger.info("子单号工作表处理完成")
    except Exception as e:
        self.logger.error(f"处理子单号工作表时出错: {str(e)}")
        raise

  def get_field_value_from_data_source(self, data_source: pd.DataFrame, data_field: str, data_field_value: str):
    """
    从数据源中获取指定列字段值等于指定值的行数据
    """
    return data_source[data_source[data_field] == data_field_value]


  def fill_sub_order_sheet(self, sub_order_sheet: Worksheet, original_detail_file_data: pd.DataFrame, original_file_data: pd.DataFrame, first_empty_row: int):
    """
    填充子单号工作表
    根据字段映射从不同的数据源获取数据：
    - detail源：original_detail_file_data（明细表数据）
    - list源：original_file_data（列表数据）
    """
    try:
        # 获取数据字段到模板字段的映射关系
        field_mappings = self.sheet_mappings["子单号"]

        # 获取模板表头到列号的映射
        header_column_mapping = self.get_header_column_mapping(sub_order_sheet)

        # 准备数据源字典
        data_sources = {
            "detail": original_detail_file_data,
            "list": original_file_data
        }

        # 检查数据源可用性
        for source_name, data_source in data_sources.items():
            if data_source is not None:
                available_columns = data_source.columns.tolist()
                self.logger.info(f"{source_name}数据源可用列: {available_columns}")
            else:
                self.logger.warning(f"{source_name}数据源为空")

        self.logger.info(f"模板表头到列号映射: {header_column_mapping}")

        # 数据行数（以明细表为准）
        data_row_count = len(original_detail_file_data)
        self.logger.info(f"需要填充 {data_row_count} 行数据")

        # 逐行填充数据
        for data_row_idx in range(data_row_count):
            target_row = first_empty_row + data_row_idx
            self.logger.debug(f"开始填充第 {data_row_idx + 1} 行数据到模板第 {target_row} 行")

            # 遍历不同数据源的字段映射关系
            for source_type, source_mappings in field_mappings.items():
                # 获取对应的数据源
                current_data_source = data_sources.get(source_type)

                if current_data_source is None:
                    self.logger.warning(f"数据源 '{source_type}' 为空，跳过")
                    continue

                # 获取当前行数据
                if source_type == "detail":
                    # detail源：直接按行索引获取
                    if data_row_idx >= len(current_data_source):
                        self.logger.warning(f"数据源 '{source_type}' 行数不足，跳过第{data_row_idx + 1}行")
                        continue
                    row_data = current_data_source.iloc[data_row_idx]
                    available_columns = current_data_source.columns.tolist()

                elif source_type == "list":
                    # list源：根据客户单号匹配
                    detail_row_data = original_detail_file_data.iloc[data_row_idx]

                    # 检查明细表是否有"客户单号"字段
                    if "客户单号" not in original_detail_file_data.columns:
                        self.logger.error("明细表中未找到'客户单号'字段，无法进行匹配")
                        continue

                    customer_order_no = detail_row_data["客户单号"]
                    self.logger.debug(f"查找客户单号: {customer_order_no}")

                    # 在list数据源中查找匹配的行
                    matched_rows = self.get_field_value_from_data_source(
                        current_data_source, "客户单号", customer_order_no
                    )

                    if matched_rows.empty:
                        self.logger.warning(f"在list数据源中未找到客户单号'{customer_order_no}'的匹配行，跳过")
                        continue
                    elif len(matched_rows) > 1:
                        self.logger.warning(f"在list数据源中找到多个客户单号'{customer_order_no}'的匹配行，使用第一个")

                    # 使用第一个匹配行的数据
                    row_data = matched_rows.iloc[0]
                    available_columns = current_data_source.columns.tolist()

                else:
                    self.logger.warning(f"未知的数据源类型: {source_type}")
                    continue

                self.logger.debug(f"处理数据源 '{source_type}' 的字段映射")

                # 遍历该数据源的字段映射关系
                for data_field, template_field in source_mappings.items():
                    try:
                        # 检查数据中是否有该字段
                        if data_field in available_columns:
                            # 检查模板中是否有对应的表头列
                            if template_field in header_column_mapping:
                                col_num = header_column_mapping[template_field]
                                field_value = row_data[data_field]

                                # 确保特定字段以文本格式填充
                                if data_field in ["子转单号", "转单号", "客户单号"]:
                                    field_value = str(field_value)

                                # 填充数据到指定位置
                                sub_order_sheet.cell(row=target_row, column=col_num).value = field_value
                                self.logger.debug(f"  填充[{source_type}]: {data_field}({field_value}) -> {template_field}(列{col_num})")
                            else:
                                self.logger.warning(f"  模板中未找到表头 '{template_field}'，跳过字段 '{data_field}'")
                        else:
                            self.logger.warning(f"  数据源'{source_type}'中未找到字段 '{data_field}'，跳过")

                    except Exception as e:
                        self.logger.error(f"  填充字段 '{data_field}' 时出错: {str(e)}")
                        continue

        self.logger.info(f"子单号数据填充完成，共填充 {data_row_count} 行")

    except Exception as e:
        self.logger.error(f"填充子单号工作表时出错: {str(e)}")
        raise

  def _find_first_empty_row(self, worksheet, max_row, max_col):
    """
    查找工作表中第一个完全空白的行

    Args:
        worksheet: openpyxl工作表对象
        max_row: 最大行数
        max_col: 最大列数

    Returns:
        int: 第一个空行的行号（从1开始）
    """
    try:
        # 从第1行开始检查，跳过可能的表头
        for row in range(1, max_row + 2):  # +2确保检查到最后一行之后
            is_empty_row = True

            # 检查该行的所有列是否都为空
            for col in range(1, max_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    is_empty_row = False
                    break

            if is_empty_row:
                self.logger.debug(f"找到第一个空行: 第{row}行")
                return row

        # 如果没有找到空行，返回最后一行的下一行
        next_row = max_row + 1
        self.logger.debug(f"未找到空行，返回下一个可用行: 第{next_row}行")
        return next_row

    except Exception as e:
        self.logger.error(f"查找空行时出错: {str(e)}")
        return max_row + 1  # 返回一个安全的行号

  def _find_collection_total_row(self, worksheet: Worksheet, max_row: int, max_col: int, search_terms: list = None, total_column: int = 1):
    """
    查找第一列中包含"Collection Total"的行

    Args:
        worksheet: openpyxl工作表对象
        max_row: 最大行数
        max_col: 最大列数
        search_terms: 搜索关键词列表
    Returns:
        int: "Collection Total"所在的行号（从1开始），未找到返回None
    """
    search_terms = search_terms or [
        "Collection Total",
        "collection total",
        "COLLECTION TOTAL",
        "Collection total",
        '合计 Total',
        "汇总",
        "总计",
        "Total"
    ]
    try:
        # 在第一列中搜索"Collection Total"
        for row in range(1, max_row + 1):
            cell_value = worksheet.cell(row=row, column=total_column).value
            if cell_value is not None:
                cell_str = str(cell_value).strip()

                for term in search_terms:
                    if term in cell_str or cell_str in term:
                        self.logger.debug(f"找到Collection Total在第{row}行: '{cell_str}'")
                        return row

        self.logger.debug("未找到包含'Collection Total'的行")
        return None

    except Exception as e:
        self.logger.error(f"查找Collection Total行时出错: {str(e)}")
        return None

  def classify_countries_data(self, original_file_data: pd.DataFrame):
    """
    按国家代码分类数据
    
    Args:
        original_file_data: 原始数据
        
    Returns:
        dict: 按国家代码分类的数据字典
    """
    try:
        summary_config = self.sheet_mappings["总结单"]
        supported_countries = summary_config["supported_countries"]
        de_postcodes = summary_config["de_postcodes"]
        
        # 检查是否有国家二字码字段
        if "国家二字码" not in original_file_data.columns:
            self.logger.error("原始数据中未找到'国家二字码'字段")
            return {}
        
        classified_data = {}
        
        # 处理DE国家特殊逻辑
        de_data = original_file_data[original_file_data["国家二字码"] == "DE"]
        if not de_data.empty:
            # DE国家中指定邮编的数据
            if "收件人邮编" in de_data.columns:
                de_specified_postcodes = de_data[de_data["收件人邮编"].astype(str).isin([str(pc) for pc in de_postcodes])]
                classified_data["DE"] = de_specified_postcodes
                self.logger.debug(f"DE指定邮编: {len(de_specified_postcodes)} 条记录")
                
                # DE国家中非指定邮编的数据归入other
                de_other_postcodes = de_data[~de_data["收件人邮编"].astype(str).isin([str(pc) for pc in de_postcodes])]
                self.logger.debug(f"DE其他邮编: {len(de_other_postcodes)} 条记录")
            else:
                # 如果没有邮编字段，所有DE数据都归入other
                de_other_postcodes = de_data
                classified_data["DE"] = pd.DataFrame()  # 空的DE数据
                self.logger.debug(f"DE数据（无邮编字段）: {len(de_other_postcodes)} 条记录")
        else:
            de_other_postcodes = pd.DataFrame()
            classified_data["DE"] = pd.DataFrame()
        
        # 分类其他支持的国家（除DE外）
        other_supported_countries = [c for c in supported_countries if c != "DE"]
        for country in other_supported_countries:
            country_data = original_file_data[original_file_data["国家二字码"] == country]
            classified_data[country] = country_data
            self.logger.debug(f"国家 {country}: {len(country_data)} 条记录")
        
        # 分类其他国家 + DE中非指定邮编的数据
        other_countries_data = original_file_data[~original_file_data["国家二字码"].isin(supported_countries)]
        
        # 合并DE其他邮编数据和其他国家数据
        if not de_other_postcodes.empty and not other_countries_data.empty:
            other_data = pd.concat([other_countries_data, de_other_postcodes], ignore_index=True)
        elif not de_other_postcodes.empty:
            other_data = de_other_postcodes
        elif not other_countries_data.empty:
            other_data = other_countries_data
        else:
            other_data = pd.DataFrame()
            
        classified_data["other"] = other_data
        self.logger.debug(f"其他类别: {len(other_data)} 条记录")
        
        self.logger.info(f"数据分类完成，共分为 {len(classified_data)} 个类别")
        return classified_data
        
    except Exception as e:
        self.logger.error(f"分类国家数据时出错: {str(e)}")
        return {}

  def count_de_postcodes(self, de_data: pd.DataFrame):
    """
    统计DE国家指定邮编的件数
    
    Args:
        de_data: DE国家的数据
        
    Returns:
        dict: 邮编对应的件数字典
    """
    try:
        summary_config = self.sheet_mappings["总结单"]
        de_postcodes = summary_config["de_postcodes"]
        
        # 检查是否有必要字段
        if "收件人邮编" not in de_data.columns:
            self.logger.error("DE数据中未找到'收件人邮编'字段")
            return {}
            
        if "件数" not in de_data.columns:
            self.logger.error("DE数据中未找到'件数'字段")
            return {}
        
        postcode_counts = {}
        
        # 统计每个指定邮编的件数
        for postcode in de_postcodes:
            # 将邮编转换为字符串进行比较
            postcode_data = de_data[de_data["收件人邮编"].astype(str) == str(postcode)]
            total_pieces = postcode_data["件数"].sum() if not postcode_data.empty else 0
            postcode_counts[postcode] = total_pieces
            self.logger.debug(f"邮编 {postcode}: {total_pieces} 件")
        
        self.logger.info(f"DE邮编统计完成，共统计 {len(de_postcodes)} 个邮编")
        return postcode_counts
        
    except Exception as e:
        self.logger.error(f"统计DE邮编件数时出错: {str(e)}")
        return {}

  def count_country_pieces(self, country_data: pd.DataFrame):
    """
    统计指定国家数据的总件数
    
    Args:
        country_data: 国家数据
        
    Returns:
        int: 总件数
    """
    try:
        if country_data.empty:
            return 0
            
        if "件数" not in country_data.columns:
            self.logger.error("数据中未找到'件数'字段")
            return 0
        
        total_pieces = country_data["件数"].sum()
        return int(total_pieces)
        
    except Exception as e:
        self.logger.error(f"统计国家件数时出错: {str(e)}")
        return 0

  def get_template_summary_sheet(self, template_workbook: Workbook):
    """
    获取模板中的总结单工作表
    
    Args:
        template_workbook: 模板工作簿
        
    Returns:
        Worksheet: 总结单工作表，失败返回None
    """
    try:
        self.logger.info("开始获取模板总结单工作表")
        sheet_names = template_workbook.sheetnames
        self.logger.info(f"模板包含的工作表: {sheet_names}")
        
        # 尝试多种可能的总结单工作表名称
        possible_names = ["总结单", "Summary", "汇总", "总结", "Summary Sheet"]
        summary_sheet = None
        for name in possible_names:
            if name in sheet_names:
                summary_sheet = template_workbook[name]
                self.logger.info(f"找到总结单工作表: '{name}'")
                break
                
        if summary_sheet is None:
            self.logger.error("未找到总结单工作表")
            return None
            
        max_row = summary_sheet.max_row
        max_col = summary_sheet.max_column
        self.logger.info(f"总结单工作表范围: {max_row}行 x {max_col}列")
        
        self.logger.info(f"总结单工作表分析完成:")
        self.logger.info(f"  - 工作表名称: '{summary_sheet.title}'")
        
        return summary_sheet
        
    except Exception as e:
        self.logger.error(f"获取总结单工作表时出错: {str(e)}")
        return None

  def process_summary_sheet(self, template_workbook: Workbook, summary_sheet: Worksheet, original_file_data: pd.DataFrame):
    """
    处理总结单工作表
    
    Args:
        template_workbook: 模板工作簿
        summary_sheet: 总结单工作表
        original_file_data: 原始数据
    """
    if template_workbook is None or summary_sheet is None:
        self.logger.error("获取总结单工作表失败")
        return False
        
    try:
        self.fill_summary_sheet(summary_sheet, original_file_data)
        self.logger.info("总结单工作表处理完成")
        return True
    except Exception as e:
        self.logger.error(f"处理总结单工作表时出错: {str(e)}")
        return False

  def fill_summary_sheet(self, summary_sheet: Worksheet, original_file_data: pd.DataFrame):
    """
    填充总结单工作表
    
    Args:
        summary_sheet: 总结单工作表
        original_file_data: 原始数据
    """
    try:
        summary_config = self.sheet_mappings["总结单"]
        data_row = summary_config["data_row"]
        
        self.logger.info("开始填充总结单数据")
        
        # 1. 按国家代码分类数据
        classified_data = self.classify_countries_data(original_file_data)
        if not classified_data:
            self.logger.error("数据分类失败，无法继续填充总结单")
            return
            
        # 2. 填充DE邮编统计
        self._fill_de_postcode_stats(summary_sheet, classified_data.get("DE", pd.DataFrame()), summary_config)
        
        # 3. 填充其他国家统计  
        self._fill_other_countries_stats(summary_sheet, classified_data, summary_config)
        
        # 4. 填充Other类别统计
        self._fill_other_category_stats(summary_sheet, classified_data.get("other", pd.DataFrame()), summary_config)
        
        # 5. 计算并填充总计
        self._fill_total_stats(summary_sheet, classified_data, summary_config)
        
        self.logger.info("总结单数据填充完成")
        
    except Exception as e:
        self.logger.error(f"填充总结单工作表时出错: {str(e)}")
        raise

  def _fill_de_postcode_stats(self, summary_sheet: Worksheet, de_data: pd.DataFrame, summary_config: dict):
    """
    填充DE邮编统计数据
    """
    try:
        if de_data.empty:
            self.logger.info("没有DE国家数据，跳过DE邮编统计")
            return
            
        self.logger.info("开始填充DE邮编统计")
        
        # 统计DE邮编件数
        postcode_counts = self.count_de_postcodes(de_data)
        
        # 填充到工作表
        data_row = summary_config["data_row"]
        start_col = summary_config["de_start_col"]
        
        for i, (postcode, count) in enumerate(postcode_counts.items()):
            col_num = start_col + i
            summary_sheet.cell(row=data_row, column=col_num).value = count
            self.logger.debug(f"填充DE邮编 {postcode}: {count} 件到列 {col_num}")
            
        self.logger.info(f"DE邮编统计填充完成，共填充 {len(postcode_counts)} 个邮编")
        
    except Exception as e:
        self.logger.error(f"填充DE邮编统计时出错: {str(e)}")
        raise

  def _fill_other_countries_stats(self, summary_sheet: Worksheet, classified_data: dict, summary_config: dict):
    """
    填充其他国家统计数据（FR、IT、ES、NL、PL、CZ、BE）
    """
    try:
        self.logger.info("开始填充其他国家统计")
        
        other_countries = ["FR", "IT", "ES", "NL", "PL", "CZ", "BE"]
        data_row = summary_config["data_row"]
        start_col = summary_config["other_countries_start_col"]
        
        for i, country in enumerate(other_countries):
            country_data = classified_data.get(country, pd.DataFrame())
            count = self.count_country_pieces(country_data)
            
            col_num = start_col + i
            summary_sheet.cell(row=data_row, column=col_num).value = count
            self.logger.debug(f"填充国家 {country}: {count} 件到列 {col_num}")
            
        self.logger.info(f"其他国家统计填充完成，共填充 {len(other_countries)} 个国家")
        
    except Exception as e:
        self.logger.error(f"填充其他国家统计时出错: {str(e)}")
        raise

  def _fill_other_category_stats(self, summary_sheet: Worksheet, other_data: pd.DataFrame, summary_config: dict):
    """
    填充Other类别统计数据
    """
    try:
        self.logger.info("开始填充Other类别统计")
        
        count = self.count_country_pieces(other_data)
        
        data_row = summary_config["data_row"]
        other_col = summary_config["other_col"]
        
        summary_sheet.cell(row=data_row, column=other_col).value = count
        self.logger.debug(f"填充Other类别: {count} 件到列 {other_col}")
        
        self.logger.info(f"Other类别统计填充完成: {count} 件")
        
    except Exception as e:
        self.logger.error(f"填充Other类别统计时出错: {str(e)}")
        raise

  def _fill_total_stats(self, summary_sheet: Worksheet, classified_data: dict, summary_config: dict):
    """
    计算并填充总计统计数据
    """
    try:
        self.logger.info("开始计算并填充总计统计")
        
        total_count = 0
        
        # 统计所有分类的件数
        for category, data in classified_data.items():
            count = self.count_country_pieces(data)
            total_count += count
            self.logger.debug(f"类别 {category}: {count} 件")
            
        data_row = summary_config["data_row"]
        total_col = summary_config["total_col"]
        
        summary_sheet.cell(row=data_row, column=total_col).value = total_count
        self.logger.debug(f"填充总计: {total_count} 件到列 {total_col}")
        
        self.logger.info(f"总计统计填充完成: {total_count} 件")
        
    except Exception as e:
        self.logger.error(f"填充总计统计时出错: {str(e)}")
        raise