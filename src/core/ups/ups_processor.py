# -*- coding: utf-8 -*-
"""
UPS专用数据处理工具
负责将原始数据按照指定映射关系填充到UPS模板的各个工作表中
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import logging
import sys
from pathlib import Path
from collections import defaultdict

# 添加项目根目录到Python路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

class UPSDataProcessor:
    """UPS数据处理器"""

    def __init__(self):
        self.logger = logging.getLogger(__name__)

        # UPS模板工作表映射关系（数据字段 -> 模板字段）
        self.sheet_mappings = {
            "总结单": {
                "转单号": "Tracking Number",
                "件数": "Packages",
                "收货实重": "G.W",
                "收货材积重": "V.G",
                "收件人邮编": "ZIP code",
                "国家二字码": "country"
            },
            "运单信息": {
                "客户单号": "参考号\n（Reference NO)",
                "件数": "件数\n(PCS)",
                "收货实重": "实重\n(Kg)",
                "收货材积重": "材重\n(Kg)",
                "国家二字码": "目的地\n(Destination)",
                "转单号": "UPS主运单号\n(Tracking Number)",
                "柜号": "提单号（集装箱/空运）"
            },
            "统计": {
                "国家": "Destination",
                "件数": "Package",
                "收货实重": "G.W",
                "收货材积重": "V.W"
            },
            "子单号": {
                "客户单号": "参考号（Reference NO)",
                "子转单号": "UPS 子单号(Tracking Number)"
            }
        }

    def get_template_workbook(self, template_path: str):
        """
        获取模板工作簿
        """
        try:
            if not Path(template_path).exists():
                self.logger.error(f"模板文件不存在: {template_path}")
                return None
            return load_workbook(template_path)
        except FileNotFoundError:
            self.logger.error(f"模板文件未找到: {template_path}")
            return None
        except PermissionError:
            self.logger.error(f"模板文件访问权限不足: {template_path}")
            return None
        except Exception as e:
            self.logger.error(f"获取模板工作簿时出错: {str(e)}")
            return None

    def process_ups_data(self, original_file_data: pd.DataFrame, detail_file: str, template_path: str, output_path: str):
        """
        处理UPS数据并填充到模板中

        Args:
            original_file_data (pd.DataFrame): 原始文件数据
            detail_file (str): 明细表文件路径
            template_path (str): UPS模板路径
            output_path (str): 输出文件路径

        Returns:
            bool: 处理结果
        """
        try:
            self.logger.info("开始处理UPS数据")

            template_workbook = self.get_template_workbook(template_path)

            original_file_data_count = original_file_data.shape[0]

            summary_sheet, first_empty_row, collection_total_row = self.get_template_summary_sheet(template_workbook)
            self.process_summary_sheet(template_workbook, summary_sheet, original_file_data, first_empty_row, collection_total_row, original_file_data_count)

            waybill_sheet, first_empty_row, collection_total_row = self.get_template_waybill_sheet(template_workbook)
            self.process_waybill_sheet(template_workbook, waybill_sheet, original_file_data, first_empty_row, collection_total_row, original_file_data_count)


            # 使用workbook对象保存文件
            template_workbook.save(output_path)
            self.logger.info(f"UPS数据处理完成，输出文件: {output_path}")
            return True
        except Exception as e:
            self.logger.error(f"处理UPS数据时出错: {str(e)}")
            return False

    def process_summary_sheet(
      self, template_workbook: Workbook, summary_sheet: Worksheet, original_file_data: pd.DataFrame, first_empty_row: int, collection_total_row: int, original_file_data_count: int):
      """
      处理总结单工作表

      Args:
        template_workbook: 模板工作簿
        summary_sheet: 总结单工作表
        original_file_data: 原始数据
        first_empty_row: 第一个空行的行号
        collection_total_row: "Collection Total"所在的行号

      Returns:
        bool: 处理结果
      """

      if template_workbook is None or summary_sheet is None:
          self.logger.error("获取模板工作表失败")
          return False

      summary_sheet_count = summary_sheet.max_row

      # 处理中间的行数是否够填充数据
      need_fill_row_count = original_file_data_count - (summary_sheet_count - first_empty_row)
      if need_fill_row_count > 0:
        # 需要填充的行数
        for i in range(need_fill_row_count):
          summary_sheet.append(original_file_data.iloc[i+first_empty_row].to_dict())
        # collection_total_row = summary_sheet_count + need_fill_row_count

      self.fill_summary_sheet(summary_sheet, original_file_data, first_empty_row)

    def process_waybill_sheet(self, template_workbook: Workbook, waybill_sheet: Worksheet, original_file_data: pd.DataFrame, first_empty_row: int, collection_total_row: int, original_file_data_count: int):
      """
      处理运单信息工作表
      """
      if template_workbook is None or waybill_sheet is None:
          self.logger.error("获取模板工作表失败")
          return False

      waybill_sheet_count = waybill_sheet.max_row

      # 处理中间的行数是否够填充数据
      need_fill_row_count = original_file_data_count - (waybill_sheet_count - first_empty_row)
      if need_fill_row_count > 0:
        # 需要填充的行数
        for i in range(need_fill_row_count):
          waybill_sheet.append(original_file_data.iloc[i+first_empty_row].to_dict())
        collection_total_row = waybill_sheet_count + need_fill_row_count

      self.fill_waybill_sheet(waybill_sheet, original_file_data, first_empty_row)

    def fill_waybill_sheet(self, waybill_sheet: Worksheet, original_file_data: pd.DataFrame, first_empty_row: int):
        """
        填充运单信息工作表
        """
        try:
            # 获取数据字段到模板字段的映射关系
            field_mappings = self.sheet_mappings["运单信息"]

            # 获取模板表头到列号的映射
            header_column_mapping = self.get_header_column_mapping(waybill_sheet)

            # 检查原始数据的列名和可用数据
            available_columns = original_file_data.columns.tolist()
            self.logger.info(f"原始数据可用列: {available_columns}")
            self.logger.info(f"模板表头到列号映射: {header_column_mapping}")

            # 数据行数
            data_row_count = len(original_file_data)
            self.logger.info(f"需要填充 {data_row_count} 行数据")

            # 逐行填充数据
            for data_row_idx in range(data_row_count):
                target_row = first_empty_row + data_row_idx
                row_data = original_file_data.iloc[data_row_idx]

                self.logger.debug(f"开始填充第 {data_row_idx + 1} 行数据到模板第 {target_row} 行")

                # 遍历字段映射关系
                for data_field, template_field in field_mappings.items():
                    try:
                        # 检查数据中是否有该字段
                        # print('data_field', data_field, 'available_columns', available_columns, '===============')
                        if data_field in available_columns:
                            # 检查模板中是否有对应的表头列
                            if template_field in header_column_mapping:
                                col_num = header_column_mapping[template_field]
                                field_value = row_data[data_field]

                                # 填充数据到指定位置
                                waybill_sheet.cell(row=target_row, column=col_num).value = field_value
                                self.logger.debug(f"  填充: {data_field}({field_value}) -> {template_field}(列{col_num})")
                            else:
                                self.logger.warning(f"  模板中未找到表头 '{template_field}'，跳过字段 '{data_field}'")
                        else:
                            self.logger.warning(f"  原始数据中未找到字段 '{data_field}'，跳过")

                    except Exception as e:
                        self.logger.error(f"  填充字段 '{data_field}' 时出错: {str(e)}")
                        continue

            self.logger.info(f"运单信息数据填充完成，共填充 {data_row_count} 行")

        except Exception as e:
            self.logger.error(f"填充运单信息工作表时出错: {str(e)}")
            raise

    def get_header_column_mapping(self, worksheet, header_row=1):
        """
        获取模板表头到列号的映射关系

        Args:
            worksheet: openpyxl工作表对象
            header_row: 表头所在行号（默认第1行）

        Returns:
            dict: 表头名称到列号的映射字典
        """
        header_mapping = {}
        try:
            max_col = worksheet.max_column
            self.logger.info(f"开始分析表头，共{max_col}列")

            for col in range(1, max_col + 1):
                header_value = worksheet.cell(row=header_row, column=col).value
                if header_value is not None:
                    header_name = str(header_value).strip()
                    if header_name:  # 非空表头
                        header_mapping[header_name] = col
                        self.logger.debug(f"表头映射: '{header_name}' -> 列{col}")

            self.logger.info(f"表头映射完成，共找到{len(header_mapping)}个有效表头")
            self.logger.debug(f"表头映射详情: {header_mapping}")

        except Exception as e:
            self.logger.error(f"分析表头时出错: {str(e)}")

        return header_mapping

    def fill_summary_sheet(self, summary_sheet: Worksheet, original_file_data: pd.DataFrame, first_empty_row: int):
        """
        填充总结单工作表
        """
        try:
            # 获取数据字段到模板字段的映射关系
            field_mappings = self.sheet_mappings["总结单"]

            # 获取模板表头到列号的映射
            header_column_mapping = self.get_header_column_mapping(summary_sheet, first_empty_row - 1)
            self.logger.info(f"模板表头到列号映射: {header_column_mapping}")

            # 检查原始数据的列名和可用数据
            available_columns = original_file_data.columns.tolist()
            self.logger.info(f"原始数据可用列: {available_columns}")
            self.logger.info(f"模板表头到列号映射: {header_column_mapping}")

            # 数据行数
            data_row_count = len(original_file_data)
            self.logger.info(f"需要填充 {data_row_count} 行数据")

            # 逐行填充数据
            for data_row_idx in range(data_row_count):
                target_row = first_empty_row + data_row_idx
                row_data = original_file_data.iloc[data_row_idx]

                self.logger.debug(f"开始填充第 {data_row_idx + 1} 行数据到模板第 {target_row} 行")

                # 遍历字段映射关系
                for data_field, template_field in field_mappings.items():
                    try:
                        # 检查数据中是否有该字段
                        if data_field in available_columns:
                            # 检查模板中是否有对应的表头列
                            if template_field in header_column_mapping:
                                col_num = header_column_mapping[template_field]
                                field_value = row_data[data_field]

                                # 填充数据到指定位置
                                summary_sheet.cell(row=target_row, column=col_num).value = field_value
                                self.logger.debug(f"  填充: {data_field}({field_value}) -> {template_field}(列{col_num})")
                            else:
                                self.logger.warning(f"  模板中未找到表头 '{template_field}'，跳过字段 '{data_field}'")
                        else:
                            self.logger.warning(f"  原始数据中未找到字段 '{data_field}'，跳过")

                    except Exception as e:
                        self.logger.error(f"  填充字段 '{data_field}' 时出错: {str(e)}")
                        continue

            self.logger.info(f"总结单数据填充完成，共填充 {data_row_count} 行")

        except Exception as e:
            self.logger.error(f"填充总结单工作表时出错: {str(e)}")
            raise

    def get_template_summary_sheet(self, template_workbook: Workbook):
        """
        获取模板中的总结单工作表，并找到关键位置信息

        Args:
            template_workbook (Workbook): UPS模板工作簿对象

        Returns:
            tuple: (worksheet, first_empty_row, collection_total_row)
                - worksheet: 总结单工作表对象
                - first_empty_row: 第一个空行的行号（从1开始）
                - collection_total_row: "Collection Total"所在的行号（从1开始），未找到返回None
        """
        try:
            # 获取总结单工作表
            summary_sheet = None
            sheet_names = template_workbook.sheetnames
            self.logger.info(f"模板包含的工作表: {sheet_names}")

            # 尝试多种可能的总结单工作表名称
            possible_names = ["总结单", "Summary", "总结", "汇总单", "汇总"]
            for name in possible_names:
                if name in sheet_names:
                    summary_sheet = template_workbook[name]
                    self.logger.info(f"找到总结单工作表: '{name}'")
                    break

            if summary_sheet is None:
                # 如果没找到，使用第一个工作表
                if sheet_names:
                    summary_sheet = template_workbook[sheet_names[0]]
                    self.logger.warning(f"未找到总结单工作表，使用第一个工作表: '{sheet_names[0]}'")
                else:
                    self.logger.error("模板文件中没有任何工作表")
                    template_workbook.close()
                    return None, None, None

            # 获取工作表的实际使用范围
            max_row = summary_sheet.max_row
            max_col = summary_sheet.max_column
            self.logger.info(f"工作表范围: {max_row}行 x {max_col}列")

            # 查找第一个空行
            first_empty_row = self._find_first_empty_row(summary_sheet, max_row, max_col)

            # 查找"Collection Total"所在的行
            collection_total_row = self._find_collection_total_row(summary_sheet, max_row, max_col)

            self.logger.info(f"总结单工作表分析完成:")
            self.logger.info(f"  - 工作表名称: '{summary_sheet.title}'")
            self.logger.info(f"  - 第一个空行: 第{first_empty_row}行")
            self.logger.info(f"  - Collection Total行: 第{collection_total_row}行" if collection_total_row else "  - Collection Total行: 未找到")

            return summary_sheet, first_empty_row, collection_total_row

        except FileNotFoundError:
            self.logger.error(f"模板文件未找到: {template_workbook}")
            return None, None, None

        except PermissionError:
            self.logger.error(f"模板文件访问权限不足: {template_workbook}")
            return None, None, None

        except Exception as e:
            self.logger.error(f"获取总结单工作表时出错: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())
            return None, None, None

    def get_template_waybill_sheet(self, template_workbook: Workbook):
        """
        获取模板中的运单信息工作表

        Args:
            template_workbook: 模板工作簿

        Returns:
            tuple: (worksheet, first_empty_row, collection_total_row)
                - worksheet: 运单信息工作表对象
                - first_empty_row: 第一个空行的行号（从1开始）
                - collection_total_row: "Collection Total"所在的行号（从1开始），未找到返回None
        """
        try:
            self.logger.info(f"开始获取模板运单信息工作表")

            sheet_names = template_workbook.sheetnames
            self.logger.info(f"模板包含的工作表: {sheet_names}")

            possible_names = ["运单信息", "Waybill", "运单", "Waybill Information", "运单信息"]
            for name in possible_names:
                if name in possible_names:
                    waybill_sheet = template_workbook[name]
                    self.logger.info(f"找到运单信息工作表: '{name}'")
                    break
            if waybill_sheet is None:
                self.logger.error("未找到运单信息工作表")
                return None, None, None, None

            max_row = waybill_sheet.max_row
            max_col = waybill_sheet.max_column
            self.logger.info(f"运单信息工作表范围: {max_row}行 x {max_col}列")

            first_empty_row = self._find_first_empty_row(waybill_sheet, max_row, max_col)
            collection_total_row = self._find_collection_total_row(waybill_sheet, max_row, max_col, ["合计 Total"])

            self.logger.info(f"运单信息工作表分析完成:")
            self.logger.info(f"  - 工作表名称: '{waybill_sheet.title}'")
            self.logger.info(f"  - 第一个空行: 第{first_empty_row}行")
            self.logger.info(f"  - Collection Total行: 第{collection_total_row}行" if collection_total_row else "  - Collection Total行: 未找到")

            return waybill_sheet, first_empty_row, collection_total_row

        except Exception as e:
            self.logger.error(f"获取运单信息工作表时出错: {str(e)}")
            return None, None, None

    def _find_first_empty_row(self, worksheet, max_row, max_col):
        """
        查找工作表中第一个完全空白的行

        Args:
            worksheet: openpyxl工作表对象
            max_row: 最大行数
            max_col: 最大列数

        Returns:
            int: 第一个空行的行号（从1开始）
        """
        try:
            # 从第1行开始检查，跳过可能的表头
            for row in range(1, max_row + 2):  # +2确保检查到最后一行之后
                is_empty_row = True

                # 检查该行的所有列是否都为空
                for col in range(1, max_col + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value is not None and str(cell_value).strip() != "":
                        is_empty_row = False
                        break

                if is_empty_row:
                    self.logger.debug(f"找到第一个空行: 第{row}行")
                    return row

            # 如果没有找到空行，返回最后一行的下一行
            next_row = max_row + 1
            self.logger.debug(f"未找到空行，返回下一个可用行: 第{next_row}行")
            return next_row

        except Exception as e:
            self.logger.error(f"查找空行时出错: {str(e)}")
            return max_row + 1  # 返回一个安全的行号

    def _find_collection_total_row(self, worksheet: Worksheet, max_row: int, max_col: int, search_terms: list = None):
        """
        查找第一列中包含"Collection Total"的行

        Args:
            worksheet: openpyxl工作表对象
            max_row: 最大行数
            max_col: 最大列数
            search_terms: 搜索关键词列表
        Returns:
            int: "Collection Total"所在的行号（从1开始），未找到返回None
        """
        search_terms = search_terms or [
            "Collection Total",
            "collection total",
            "COLLECTION TOTAL",
            "Collection total",
            '合计 Total',
            "汇总",
            "总计"
        ]
        try:
            # 在第一列中搜索"Collection Total"
            for row in range(1, max_row + 1):
                cell_value = worksheet.cell(row=row, column=1).value
                if cell_value is not None:
                    cell_str = str(cell_value).strip()



                    for term in search_terms:
                        if term in cell_str or cell_str in term:
                            self.logger.debug(f"找到Collection Total在第{row}行: '{cell_str}'")
                            return row

            self.logger.debug("未找到包含'Collection Total'的行")
            return None

        except Exception as e:
            self.logger.error(f"查找Collection Total行时出错: {str(e)}")
            return None
