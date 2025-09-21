# -*- coding: utf-8 -*-
"""
模板数据填充功能
负责将Excel数据填充到指定模板并生成输出文件
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import logging
import sys
from pathlib import Path
from datetime import datetime
import shutil

# 添加项目根目录到Python路径
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from config import *
from src.core.ups_processor import UPSDataProcessor

class TemplateFiller:
    """模板数据填充器"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.ups_processor = UPSDataProcessor()
        
    def fill_template(self, data, template_path, template_type):
        """
        填充模板数据
        
        Args:
            data (dict): Excel数据字典
            template_path (str): 模板文件路径
            template_type (str): 模板类型 ("UPS" 或 "DPD")
            
        Returns:
            str: 输出文件路径，失败返回None
        """
        try:
            self.logger.info(f"开始填充模板: {template_type}")
            self.logger.info(f"模板文件: {template_path}")
            
            # 验证模板文件
            if not self.validate_template(template_path):
                return None
                
            # 生成输出文件路径
            output_path = self.generate_output_path(template_type, data)
            
            # 复制模板文件到输出路径
            shutil.copy2(template_path, output_path)
            
            # 根据模板类型选择填充方法
            if template_type == "UPS":
                success = self.fill_ups_template(data, output_path)
            elif template_type == "DPD":
                success = self.fill_dpd_template(data, output_path)
            else:
                self.logger.error(f"未知的模板类型: {template_type}")
                return None
                
            if success:
                self.logger.info(f"模板填充成功: {output_path}")
                return str(output_path)
            else:
                # 删除失败的文件
                if Path(output_path).exists():
                    Path(output_path).unlink()
                return None
                
        except Exception as e:
            self.logger.error(f"填充模板时出错: {str(e)}")
            return None
            
    def validate_template(self, template_path):
        """
        验证模板文件
        
        Args:
            template_path (str): 模板文件路径
            
        Returns:
            bool: 验证结果
        """
        try:
            # 检查文件是否存在
            if not Path(template_path).exists():
                self.logger.error(f"模板文件不存在: {template_path}")
                return False
                
            # 检查文件是否可读
            try:
                wb = load_workbook(template_path)
                wb.close()
                return True
            except Exception as e:
                self.logger.error(f"模板文件无法打开: {str(e)}")
                return False
                
        except Exception as e:
            self.logger.error(f"验证模板时出错: {str(e)}")
            return False
            
    def generate_output_path(self, template_type, data):
        """
        生成输出文件路径
        
        Args:
            template_type (str): 模板类型
            data (dict): 数据字典
            
        Returns:
            Path: 输出文件路径
        """
        try:
            # 获取当前时间戳
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 获取原始文件名（不含扩展名）
            original_name = "数据"
            if data and '_metadata' in data:
                original_name = Path(data['_metadata']['file_name']).stem
                
            # 生成输出文件名
            output_name = f"{template_type}_{original_name}_{timestamp}.xlsx"
            
            # 输出到桌面
            output_path = DESKTOP_PATH / output_name
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"生成输出路径时出错: {str(e)}")
            # 使用默认路径
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            return DESKTOP_PATH / f"{template_type}_输出_{timestamp}.xlsx"
            
    def fill_ups_template(self, data, output_path):
        """
        填充UPS总结单模板
        使用专用的UPS数据处理器
        
        Args:
            data (dict): Excel数据字典
            output_path (str): 输出文件路径
            
        Returns:
            bool: 填充结果
        """
        try:
            self.logger.info("开始使用UPS专用处理器填充模板")
            
            # 使用UPS专用处理器
            success = self.ups_processor.process_ups_data(
                source_data=data,
                template_path=output_path,  # 这里output_path实际是复制后的模板文件
                output_path=output_path
            )
            
            if success:
                self.logger.info("UPS模板填充完成")
                return True
            else:
                self.logger.error("UPS模板填充失败")
                return False
                
        except Exception as e:
            self.logger.error(f"填充UPS模板时出错: {str(e)}")
            return False
            
    def fill_dpd_template(self, data, output_path):
        """
        填充DPD数据预报模板
        
        Args:
            data (dict): Excel数据字典
            output_path (str): 输出文件路径
            
        Returns:
            bool: 填充结果
        """
        try:
            self.logger.info("开始填充DPD数据预报模板")
            
            # 打开模板文件
            wb = load_workbook(output_path)
            
            # 获取第一个工作表作为主工作表
            if wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            else:
                self.logger.error("模板文件没有工作表")
                return False
                
            # 获取源数据的第一个工作表
            source_data = None
            for sheet_name, sheet_data in data.items():
                if sheet_name != '_metadata':
                    source_data = sheet_data
                    break
                    
            if not source_data:
                self.logger.error("源数据为空")
                return False
                
            # DPD模板填充逻辑
            self.logger.info(f"源数据列: {source_data['columns']}")
            self.logger.info(f"源数据行数: {source_data['shape'][0]}")
            
            # 查找数据起始位置
            data_start_row = self.find_data_start_row(ws, ["预报", "数据", "详情"])
            
            if data_start_row:
                self.fill_data_to_worksheet(ws, source_data['data'], data_start_row)
            else:
                # 如果找不到标记，从第3行开始填充
                self.fill_data_to_worksheet(ws, source_data['data'], 3)
                
            # 填充汇总信息
            self.fill_summary_info(ws, source_data, "DPD")
            
            # 保存文件
            wb.save(output_path)
            wb.close()
            
            self.logger.info("DPD模板填充完成")
            return True
            
        except Exception as e:
            self.logger.error(f"填充DPD模板时出错: {str(e)}")
            return False
            
    def find_data_start_row(self, worksheet, keywords):
        """
        查找数据起始行
        
        Args:
            worksheet: openpyxl工作表对象
            keywords (list): 关键词列表
            
        Returns:
            int: 起始行号，找不到返回None
        """
        try:
            for row in range(1, 20):  # 搜索前20行
                for col in range(1, 10):  # 搜索前10列
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        for keyword in keywords:
                            if keyword in cell_str:
                                return row + 1  # 返回下一行作为数据起始行
                                
            return None
            
        except Exception as e:
            self.logger.error(f"查找数据起始行时出错: {str(e)}")
            return None
            
    def fill_data_to_worksheet(self, worksheet, data_records, start_row):
        """
        将数据填充到工作表
        
        Args:
            worksheet: openpyxl工作表对象
            data_records (list): 数据记录列表
            start_row (int): 起始行号
        """
        try:
            if not data_records:
                return
                
            # 获取数据列名
            columns = list(data_records[0].keys()) if data_records else []
            
            # 填充表头（如果需要）
            for col_idx, column_name in enumerate(columns, 1):
                cell = worksheet.cell(row=start_row-1, column=col_idx)
                if not cell.value:  # 只在空单元格填充表头
                    cell.value = column_name
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    
            # 填充数据
            for row_idx, record in enumerate(data_records, start_row):
                for col_idx, column_name in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    value = record.get(column_name)
                    
                    # 处理空值
                    if pd.isna(value) or value is None:
                        cell.value = ""
                    else:
                        cell.value = value
                        
            self.logger.info(f"已填充 {len(data_records)} 行数据到工作表")
            
        except Exception as e:
            self.logger.error(f"填充数据到工作表时出错: {str(e)}")
            
    def fill_summary_info(self, worksheet, source_data, template_type):
        """
        填充汇总信息
        
        Args:
            worksheet: openpyxl工作表对象
            source_data (dict): 源数据
            template_type (str): 模板类型
        """
        try:
            # 查找汇总信息位置（通常在模板顶部）
            summary_cells = [
                ("总计", len(source_data['data'])),
                ("处理时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("模板类型", template_type)
            ]
            
            # 在前几行查找并填充汇总信息
            for row in range(1, 10):
                for col in range(1, 10):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        for summary_key, summary_value in summary_cells:
                            if summary_key in cell_str:
                                # 在右侧单元格填充值
                                target_cell = worksheet.cell(row=row, column=col+1)
                                target_cell.value = summary_value
                                break
                                
        except Exception as e:
            self.logger.error(f"填充汇总信息时出错: {str(e)}")